import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('/Users/admin/receipt_app_project/Plot Fin.xlsx')
ws = wb.active

print("="*80)
print("COMPLETE EXCEL ANALYSIS")
print("="*80)

# Read all cells to understand the complete structure
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        if cell.value is not None:
            formula_info = ""
            if cell.data_type == 'f':
                formula_info = f" [FORMULA: {cell.value}]"
            
            # Check cell color
            fill = cell.fill
            color_info = ""
            if fill and fill.start_color and hasattr(fill.start_color, 'rgb'):
                rgb = fill.start_color.rgb
                if rgb and 'FFFF' in str(rgb):
                    color_info = " 🟡 [INPUT]"
            
            print(f"{cell.coordinate}: {cell.value}{formula_info}{color_info}")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)
print(f"Total rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")
